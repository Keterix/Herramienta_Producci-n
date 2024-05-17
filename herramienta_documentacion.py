import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Función para limpiar caracteres inválidos en nombres de hojas de Excel
def limpiar_nombre(nombre):
    caracteres_no_validos = ['\\', '/', '*', '[', ']', ':', '?']
    for caracter in caracteres_no_validos:
        nombre = nombre.replace(caracter, '_')
    return nombre[:30]

# Función para listar archivos dentro de las carpetas especificadas
def listar_archivos(carpetas):
    archivos = {}
    for carpeta in carpetas:
        archivos[carpeta] = []
        for ruta, _, archivos_en_carpeta in os.walk(carpeta):
            for archivo in archivos_en_carpeta:
                ruta_completa = os.path.join(ruta, archivo)
                try:
                    fecha_creacion = datetime.fromtimestamp(os.path.getctime(ruta_completa))
                    fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(ruta_completa))
                    nombre, extension = os.path.splitext(archivo)
                    ruta_padre = os.path.dirname(ruta_completa)
                    archivos[carpeta].append((nombre, extension, fecha_creacion, fecha_modificacion, ruta_padre))
                except PermissionError as e:
                    print(f"Error de permisos con el archivo: {ruta_completa} - {e}")
    return archivos

# Función para agregar una hoja al libro de Excel
def agregar_hoja_excel(libro_excel, nombre_hoja):
    if nombre_hoja in libro_excel.sheetnames:
        print(f"La hoja '{nombre_hoja}' ya existe en el archivo de Excel.")
        nuevo_nombre = input("Ingrese un nuevo nombre para la hoja: ")
        return agregar_hoja_excel(libro_excel, nuevo_nombre)
    else:
        libro_excel.create_sheet(title=nombre_hoja)
        return nombre_hoja

# Función para ajustar el ancho de las columnas
def ajustar_ancho_columnas(hoja):
    for col in hoja.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        hoja.column_dimensions[column].width = adjusted_width

# Función para guardar datos en un archivo de Excel
def guardar_en_excel(archivos, nombre_archivo, nombre_hoja_nueva=None):
    if not nombre_archivo:
        nombre_archivo = "1714416546463_Anexo Circular 02_2024.xlsx"
    
    # Asegurarse de que el nombre del archivo tenga la extensión .xlsx
    if not nombre_archivo.lower().endswith('.xlsx'):
        nombre_archivo += '.xlsx'
    
    if os.path.exists(nombre_archivo):
        libro_excel = load_workbook(nombre_archivo)
    else:
        libro_excel = Workbook()
    
    for carpeta, lista_archivos in archivos.items():
        nombre_hoja = limpiar_nombre(carpeta)
        if nombre_hoja_nueva:
            nombre_hoja_final = agregar_hoja_excel(libro_excel, nombre_hoja_nueva)
        else:
            nombre_hoja_final = agregar_hoja_excel(libro_excel, nombre_hoja)
        
        hoja = libro_excel[nombre_hoja_final]
        
        hoja.append(["ID", "Nombre de dato / archivo", "Formato", "Fecha de creación", "Fecha de modificación", "Ruta",
                     "Responsable del dato / archivo", "Propósito del dato / archivo",
                     "¿Quién tiene acceso al archivo / dato?", "Transferencia con 3ero / externos",
                     "Responsable de respaldo", "Fecha de respaldo", "Responsable de eliminación", "Fecha de eliminación"])
        
        for id_archivo, archivo in enumerate(lista_archivos, start=1):
            hoja.append([id_archivo] + list(archivo))
        
        ajustar_ancho_columnas(hoja)  # Ajustar el ancho de las columnas después de llenar la hoja
    
    try:
        libro_excel.save(nombre_archivo)
    except PermissionError as e:
        print(f"Error de permisos al guardar el archivo: {nombre_archivo} - {e}")

# Función principal
def main():
    print("La ruta del disco puede ser solo la letra que lo representa como 'E', 'D', 'C', con su respectivo :\\")
    ruta_disco = input("Ingrese la ruta del disco a analizar (por ejemplo, C:\\): ")
    
    # Asegurar que la ruta tenga dobles barras invertidas
    ruta_disco = ruta_disco.replace("\\", "\\\\")
    
    if not os.path.exists(ruta_disco):
        print("La ruta del disco especificada no es válida.")
        return

    carpetas = [ruta_disco]
    archivos = listar_archivos(carpetas)
    print('')
    print("El archivo por defecto es '1714416546463_Anexo Circular 02_2024.xlsx'")
    print('No es necesario agregar la extension xlsx')
    nombre_archivo_excel = input("Ingrese el nombre del archivo de Excel donde se guardarán los datos (o presione Enter para usar el nombre por defecto): ")
    if not nombre_archivo_excel:
        nombre_archivo_excel = "1714416546463_Anexo Circular 02_2024.xlsx"
    
    # Asegurarse de que el nombre del archivo tenga la extensión .xlsx
    if not nombre_archivo_excel.lower().endswith('.xlsx'):
        nombre_archivo_excel += '.xlsx'
    
    print('')
    nombre_hoja_nueva = input("Ingrese el nombre para la nueva hoja (o presione Enter para usar la ruta del disco como nombre de la hoja): ")
    if not nombre_hoja_nueva:
        nombre_hoja_nueva = limpiar_nombre(ruta_disco.strip(os.sep))
    
    guardar_en_excel(archivos, nombre_archivo_excel, nombre_hoja_nueva)

if __name__ == "__main__":
    main()
